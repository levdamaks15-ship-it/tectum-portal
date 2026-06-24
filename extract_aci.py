import openpyxl
import sys

file_path = "docs/excel/рапорт_АЦИ 10.06.26..xlsx"
with open("aci_info.txt", "w", encoding="utf-8") as f:
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            f.write(f"\n--- Sheet: {sheet_name} ---\n")
            sheet = wb[sheet_name]
            for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True)):
                if any(cell is not None for cell in row):
                    f.write(f"Row {idx+1}: {row}\n")
    except Exception as e:
        f.write(f"Error: {e}\n")
