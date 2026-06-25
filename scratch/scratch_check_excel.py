import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook("monthly_plan_board.xlsx", data_only=True)
print("Sheetnames:", wb.sheetnames)
ws = wb.active
print("Active sheet title:", ws.title)
for r in list(ws.iter_rows(values_only=True))[:15]:
    print(r)
