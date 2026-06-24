import openpyxl
import sys

file_path = "docs/excel/рапорт_АЦИ 10.06.26..xlsx"
with open("aci_full_headers.txt", "w", encoding="utf-8") as f:
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb['рапорт']
        
        # Get dimensions
        max_col = sheet.max_column
        max_row = sheet.max_row
        
        f.write(f"Max cols: {max_col}, Max rows: {max_row}\n\n")
        
        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True)):
            f.write(f"Row {idx+1} length {len(row)}:\n")
            for i, val in enumerate(row):
                f.write(f"  Col {i+1}: {val}\n")
    except Exception as e:
        f.write(f"Error: {e}\n")
