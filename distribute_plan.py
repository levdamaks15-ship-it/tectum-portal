import openpyxl
from datetime import datetime, timedelta

def generate_plan(month=6, year=2026, target=160000):
    # Determine the number of days in the month
    if month in [4, 6, 9, 11]:
        days_in_month = 30
    elif month == 2:
        days_in_month = 29 if year % 4 == 0 else 28
    else:
        days_in_month = 31

    # Figure out the exact mondays
    mondays = []
    for d in range(1, days_in_month + 1):
        dt = datetime(year, month, d)
        if dt.weekday() == 0: # 0 is Monday
            mondays.append(d)

    # Weights
    weights = []
    shifts = []
    
    for d in range(1, days_in_month + 1):
        # Day Shift
        w_day = 5.5 if d in mondays else 11.0
        weights.append(w_day)
        shifts.append((d, "День", w_day))
        
        # Night Shift
        w_night = 13.0
        weights.append(w_night)
        shifts.append((d, "Ночь", w_night))

    total_weight = sum(weights)
    exact_values = [w / total_weight * target for w in weights]
    
    plans = []
    running_exact = 0.0
    running_int = 0
    for ev in exact_values:
        running_exact += ev
        val = round(running_exact) - running_int
        plans.append(val)
        running_int += val

    # Read existing facts if the file exists
    facts = {}
    try:
        wb_old = openpyxl.load_workbook("monthly_plan_board.xlsx", data_only=True)
        ws_old = wb_old.active
        for row in ws_old.iter_rows(min_row=2, values_only=True):
            if len(row) >= 6:
                date_str, shift_type, _, _, _, fact, *_ = row
                if date_str and shift_type:
                    facts[(date_str, shift_type)] = fact
    except Exception as e:
        print(f"Could not read old facts: {e}")

    # Masters and pattern
    # Sequence of masters for days:
    # 01.06 Day=3, Night=1
    # 02.06 Day=4, Night=3
    # 03.06 Day=2, Night=4
    # 04.06 Day=1, Night=2
    # So the Day sequence is 3, 4, 2, 1 repeating.
    # The Night sequence is 1, 3, 4, 2 repeating.
    
    day_masters = [3, 4, 2, 1]
    night_masters = [1, 3, 4, 2]
    
    master_names = {
        1: "Бекбосынов Р.",
        2: "Монаев С.",
        3: "Султанулы С.",
        4: "Дауылбай М."
    }

    # Determine starting index for June 1st
    # Let's say we anchor to June 1st, 2026 as index 0.
    anchor_date = datetime(2026, 6, 1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Выработка"
    headers = ["Дата", "Тип смены", "Мастер", "Смена", "План", "Факт"]
    ws.append(headers)

    for i, (d, shift_type, w) in enumerate(shifts):
        dt = datetime(year, month, d)
        delta_days = (dt - anchor_date).days
        
        idx = delta_days % 4
        if shift_type == "День":
            master_num = day_masters[idx]
        else:
            master_num = night_masters[idx]
            
        master_name = master_names[master_num]
        date_str = dt.strftime("%d.%m.%Y")
        plan_val = plans[i]
        
        fact_val = facts.get((date_str, shift_type), "")
        
        ws.append([date_str, shift_type, master_name, master_num, plan_val, fact_val])

    wb.save("monthly_plan_board_v2.xlsx")
    print(f"Generated monthly_plan_board_v2.xlsx with total plan: {sum(plans)} sheets.")

if __name__ == '__main__':
    generate_plan(6, 2026, 160000)
