import urllib.request

url = "http://localhost:8000/api/dashboard/export_daily_report?start_date=2026-06-10&line=lfm2"
try:
    with urllib.request.urlopen(url) as response:
        content = response.read()
        print("Status code:", response.status)
        print("Headers:")
        for k, v in response.info().items():
            print(f"  {k}: {v}")
        print("Content length:", len(content))
        with open("test_report.xlsx", "wb") as f:
            f.write(content)
        print("Saved test_report.xlsx")
        
        import openpyxl
        wb = openpyxl.load_workbook("test_report.xlsx")
        print("Sheets in downloaded file:", wb.sheetnames)
        for name in wb.sheetnames:
            print(f"  Name: {name!r}, bytes: {name.encode('utf-8')}")
except Exception as e:
    print("Error:", e)

